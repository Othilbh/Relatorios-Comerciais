# utils/builder.py
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.consolidacao import calcular_margem, consolidar_clientes, normalizar_cliente
from utils.categorias import categorizar_produto, ORDEM_COLUNAS, margem_real

COR_HEADER    = "1A3A5C"
COR_SUBHEADER = "2E6DA4"
COR_ACCENT    = "F5A623"
COR_VERDE     = "C6EFCE"
COR_VERDE_FONT= "276221"
COR_VERMELHO  = "FFC7CE"
COR_VERMELHO_FONT = "9C0006"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size)

def _border_thin():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _alinhar(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_cell(ws, row, col, valor, bg=COR_HEADER, fg="FFFFFF", bold=True, size=10, align="center"):
    c = ws.cell(row=row, column=col, value=valor)
    c.fill = _fill(bg)
    c.font = _font(bold=bold, color=fg, size=size)
    c.alignment = _alinhar(align)
    c.border = _border_thin()
    return c

def _data_cell(ws, row, col, valor, fmt=None, bg=None, bold=False, align="center"):
    c = ws.cell(row=row, column=col, value=valor)
    if fmt:
        c.number_format = fmt
    if bg:
        c.fill = _fill(bg)
    c.font = _font(bold=bold)
    c.alignment = _alinhar(align)
    c.border = _border_thin()
    return c

def _margem_cell(ws, row, col, margem):
    bg = COR_VERDE if margem >= 0 else COR_VERMELHO
    fg = COR_VERDE_FONT if margem >= 0 else COR_VERMELHO_FONT
    c = ws.cell(row=row, column=col, value=round(margem, 2))
    c.number_format = '0.00"%"'
    c.fill = _fill(bg)
    c.font = _font(color=fg)
    c.alignment = _alinhar("center")
    c.border = _border_thin()
    return c

def criar_aba_leiame(wb, data_ref):
    ws = wb.create_sheet("Leia-me")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 80
    ws.merge_cells("A1:A2")
    c = ws["A1"]
    c.value = "📊 RELATÓRIO COMERCIAL OTHIL"
    c.font = Font(bold=True, size=16, color=COR_HEADER)
    c.alignment = _alinhar("center")
    infos = [
        ("", ""),
        ("Data de referência:", data_ref),
        ("Gerado por:", "Sistema de Relatórios OTHIL"),
        ("", ""),
        ("REGRA DE MARGEM:", "Resultado % = (Faturamento - Custo Total) / Faturamento × 100"),
        ("", "Calculado produto a produto. NUNCA usar % do rodapé do PDF."),
        ("Margem Real:", "Margem Rel + 15%"),
    ]
    for i, (chave, valor) in enumerate(infos, start=3):
        ws.cell(row=i, column=1, value=f"  {chave}  {valor}" if chave else "")
    return ws

def criar_aba_vendedor(wb, nome_vendedor, dados_clientes):
    ws = wb.create_sheet(nome_vendedor[:31])
    ws.sheet_view.showGridLines = False
    consolidado = consolidar_clientes(dados_clientes)

    _header_cell(ws, 1, 1, f"PRODUTOS — {nome_vendedor.upper()}", bg=COR_HEADER, size=12, align="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    for col, h in enumerate(["Cliente", "Produto", "Qtd CX", "Fat R$", "Custo R$", "Margem %"], 1):
        _header_cell(ws, 2, col, h, bg=COR_SUBHEADER)

    row = 3
    for item in dados_clientes:
        nome_cli = normalizar_cliente(item["cliente"])
        for prod in item.get("produtos", []):
            fat_p = prod.get("faturamento", 0)
            custo_p = prod.get("custo_total", prod.get("qtd", 0) * prod.get("custo_unit", 0))
            _data_cell(ws, row, 1, nome_cli, align="left")
            _data_cell(ws, row, 2, prod.get("descricao", ""), align="left")
            _data_cell(ws, row, 3, prod.get("qtd", 0), fmt='#,##0.0')
            _data_cell(ws, row, 4, fat_p, fmt='R$ #,##0.00')
            _data_cell(ws, row, 5, custo_p, fmt='R$ #,##0.00')
            _margem_cell(ws, row, 6, calcular_margem(fat_p, custo_p))
            row += 1

    _header_cell(ws, 1, 8, f"RESUMO — {nome_vendedor.upper()}", bg=COR_ACCENT, fg="000000", size=12, align="left")
    ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=12)
    for col, h in enumerate(["Cliente", "Vol CX", "Fat R$", "Custo R$", "Margem %"], 8):
        _header_cell(ws, 2, col, h, bg=COR_SUBHEADER)

    row_r = 3
    total_fat = total_custo = total_vol = 0.0
    for nome_cli, vals in sorted(consolidado.items()):
        fat, custo, vol = vals["faturamento"], vals["custo"], vals["volume"]
        total_fat += fat; total_custo += custo; total_vol += vol
        _data_cell(ws, row_r, 8, nome_cli, align="left")
        _data_cell(ws, row_r, 9, vol, fmt='#,##0.0')
        _data_cell(ws, row_r, 10, fat, fmt='R$ #,##0.00')
        _data_cell(ws, row_r, 11, custo, fmt='R$ #,##0.00')
        _margem_cell(ws, row_r, 12, calcular_margem(fat, custo))
        row_r += 1

    _header_cell(ws, row_r, 8, "TOTAL", bg=COR_HEADER)
    _data_cell(ws, row_r, 9, total_vol, fmt='#,##0.0', bold=True)
    _data_cell(ws, row_r, 10, total_fat, fmt='R$ #,##0.00', bold=True)
    _data_cell(ws, row_r, 11, total_custo, fmt='R$ #,##0.00', bold=True)
    _margem_cell(ws, row_r, 12, calcular_margem(total_fat, total_custo))

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32
    for col in ["C","D","E","F"]: ws.column_dimensions[col].width = 13
    ws.column_dimensions["H"].width = 28
    for col in ["I","J","K","L"]: ws.column_dimensions[col].width = 14

    return ws, {"faturamento": total_fat, "custo": total_custo, "volume": total_vol}

def criar_aba_consolidado(wb, resumo_vendedores):
    ws = wb.create_sheet("Consolidado", 1)
    ws.sheet_view.showGridLines = False
    _header_cell(ws, 1, 1, "CONSOLIDADO GERAL — OTHIL", bg=COR_HEADER, size=13)
    ws.merge_cells("A1:E1")
    for col, h in enumerate(["Vendedor","Vol CX","Fat R$","Custo R$","Margem %"], 1):
        _header_cell(ws, 2, col, h, bg=COR_SUBHEADER)

    row = 3
    total_fat = total_custo = total_vol = 0.0
    for vendedor, vals in sorted(resumo_vendedores.items()):
        fat, custo, vol = vals["faturamento"], vals["custo"], vals["volume"]
        total_fat += fat; total_custo += custo; total_vol += vol
        _data_cell(ws, row, 1, vendedor, align="left")
        _data_cell(ws, row, 2, vol, fmt='#,##0.0')
        _data_cell(ws, row, 3, fat, fmt='R$ #,##0.00')
        _data_cell(ws, row, 4, custo, fmt='R$ #,##0.00')
        _margem_cell(ws, row, 5, calcular_margem(fat, custo))
        row += 1

    _header_cell(ws, row, 1, "TOTAL GERAL", bg=COR_HEADER)
    _data_cell(ws, row, 2, total_vol, fmt='#,##0.0', bold=True)
    _data_cell(ws, row, 3, total_fat, fmt='R$ #,##0.00', bold=True)
    _data_cell(ws, row, 4, total_custo, fmt='R$ #,##0.00', bold=True)
    _margem_cell(ws, row, 5, calcular_margem(total_fat, total_custo))

    ws.column_dimensions["A"].width = 20
    for col in ["B","C","D","E"]: ws.column_dimensions[col].width = 15
    return ws

def criar_aba_recorrencia(wb, todos_clientes):
    ws = wb.create_sheet("Recorrência")
    ws.sheet_view.showGridLines = False
    matriz = {}
    for item in todos_clientes:
        nome = normalizar_cliente(item["cliente"])
        if nome not in matriz:
            matriz[nome] = {}
        for prod in item.get("produtos", []):
            cat = categorizar_produto(prod.get("descricao", ""))
            if cat not in matriz[nome]:
                matriz[nome][cat] = {"qtd": 0.0, "faturamento": 0.0, "custo": 0.0}
            matriz[nome][cat]["qtd"] += prod.get("qtd", 0)
            matriz[nome][cat]["faturamento"] += prod.get("faturamento", 0)
            custo = prod.get("custo_total", prod.get("qtd", 0) * prod.get("custo_unit", 0))
            matriz[nome][cat]["custo"] += custo

    cats_usadas = [c for c in ORDEM_COLUNAS if any(c in m for m in matriz.values())]
    if any("OUTROS" in m for m in matriz.values()):
        cats_usadas.append("OUTROS")

    _header_cell(ws, 1, 1, "MATRIZ DE RECORRÊNCIA — OTHIL", bg=COR_HEADER, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cats_usadas)+2)
    _header_cell(ws, 2, 1, "Cliente", bg=COR_SUBHEADER)
    for col_i, cat in enumerate(cats_usadas, 2):
        _header_cell(ws, 2, col_i, cat, bg=COR_SUBHEADER, size=8)
    _header_cell(ws, 2, len(cats_usadas)+2, "Margem Real %", bg=COR_ACCENT, fg="000000")

    row = 3
    for cliente in sorted(matriz.keys()):
        ws.cell(row=row, column=1, value=cliente).alignment = _alinhar("left")
        fat_total = custo_total = 0.0
        for col_i, cat in enumerate(cats_usadas, 2):
            d = matriz[cliente].get(cat)
            if d and d["qtd"] > 0:
                fat_total += d["faturamento"]; custo_total += d["custo"]
                m_c = calcular_margem(d["faturamento"], d["custo"])
                c = ws.cell(row=row, column=col_i, value=round(d["qtd"], 1))
                c.fill = _fill(COR_VERDE if m_c >= 0 else COR_VERMELHO)
                c.font = _font(color=COR_VERDE_FONT if m_c >= 0 else COR_VERMELHO_FONT)
                c.alignment = _alinhar("center"); c.border = _border_thin()
            else:
                ws.cell(row=row, column=col_i, value="").border = _border_thin()
        mr = margem_real(calcular_margem(fat_total, custo_total))
        _margem_cell(ws, row, len(cats_usadas)+2, mr)
        row += 1

    ws.column_dimensions["A"].width = 30
    for i in range(2, len(cats_usadas)+3):
        ws.column_dimensions[get_column_letter(i)].width = 12
    return ws

def gerar_excel(lista_dados: list, data_ref: str = "") -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    criar_aba_leiame(wb, data_ref)
    resumo_vendedores = {}
    todos_clientes = []
    for dados in lista_dados:
        vendedor = dados.get("vendedor", "Desconhecido")
        clientes = dados.get("clientes", [])
        if not clientes:
            continue
        _, resumo = criar_aba_vendedor(wb, vendedor, clientes)
        resumo_vendedores[vendedor] = resumo
        todos_clientes.extend(clientes)
    if resumo_vendedores:
        criar_aba_consolidado(wb, resumo_vendedores)
    if todos_clientes:
        criar_aba_recorrencia(wb, todos_clientes)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf