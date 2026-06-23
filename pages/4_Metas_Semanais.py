# pages/4_Metas_Semanais.py
import streamlit as st
from datetime import date, timedelta
import sys, os, tempfile
import pandas as pd
import io, json, math, base64, requests
sys.path.append(os.path.dirname(os.path.dirname(__file__)))

from utils.parser import extrair_vendas_por_vendedor, extrair_estoque_por_vendedor, VENDEDORES_ATIVOS
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Metas Semanais · OTHIL", page_icon="🎯", layout="wide")
st.title("🎯 Metas Semanais")
st.caption("Relatório diário por vendedor + metas semanais. Luca excluído automaticamente.")
st.divider()

PERCENTUAIS = {
    "Farley": 0.175, "Dora": 0.175, "Afanais": 0.250, "Roni": 0.250,
    "Reginaldo": 0.225, "Juliana": 0.075, "Claudia": 0.075, "Luciano": 0.075,
}

PRODUTOS_DEFAULT = [
    "Portuguesa Sabor 55/60", "Portuguesa 60/70", "Portuguesa 70/80",
    "Forelle", "Ercoline", "Pera Asiática", "Gala Santa Carol", "Gala Azaleia",
    "Fuji Expressa 180", "Fuji Azaleia", "Fuji Hiragami", "Fuji Suprema",
    "Thompson Vitace", "Thompson Seedless", "Uva Crimson", "Uva Isis",
    "Uva Jubilee", "Uva Itália", "Maçã Argentina", "Maçã Chilena",
    "Maçã Pink Lady", "Maçã Granny Smith", "Maçã Red Globe",
    "Mamão Havai", "Mamão Formoso", "Goiaba", "Melão Amarelo", "Melão Gaia",
    "Melão Cantaloupe", "Tangerina Cumbuca", "Tangerina Ponkan",
    "Ameixa", "Pêssego", "Nectarina", "Morango", "Mirtilo", "Abacaxi",
    "Manga Palmer", "Manga Tommy", "Laranja", "Limão Tahiti", "Limão Siciliano", "Tomate Roma",
]

MAPA_PRODUTO = {
    "Portuguesa Sabor 55/60": ["PORTUGUESA SABOR", "PORTUGUESA 55", "02050032"],
    "Portuguesa 60/70": ["PORTUGUESA 60"], "Portuguesa 70/80": ["PORTUGUESA 70"],
    "Forelle": ["FORELLE", "020502701"], "Ercoline": ["ERCOLINE"],
    "Pera Asiática": ["ASIATICA", "ASIÁTICA"], "Gala Santa Carol": ["GALA SANTA", "SANTA CAROL"],
    "Gala Azaleia": ["GALA AZALEIA"], "Fuji Expressa 180": ["FUJI EXPRESSA", "FUJI 180"],
    "Fuji Azaleia": ["FUJI AZALEIA", "702145135", "702145165"], "Fuji Hiragami": ["HIRAGAMI"],
    "Fuji Suprema": ["FUJI SUPREMA"], "Thompson Vitace": ["THOMPSON VITACE"],
    "Thompson Seedless": ["THOMPSON SEEDLESS"], "Uva Crimson": ["CRINSON", "CRIMSON"],
    "Uva Isis": ["ISIS"], "Uva Jubilee": ["JUBILEE"], "Uva Itália": ["ITALIA", "ITÁLIA"],
    "Maçã Argentina": ["ARGENTINA"], "Maçã Chilena": ["CHILENA"],
    "Maçã Pink Lady": ["PINK LADY"], "Maçã Granny Smith": ["GRAN SMITH", "GRANNY SMITH"],
    "Maçã Red Globe": ["RED GLOBE"], "Mamão Havai": ["MAMAO HAVAI", "MAMÃO HAVAI"],
    "Mamão Formoso": ["MAMAO FORMOSO", "MAMÃO FORMOSO"],
    "Goiaba": ["GOIABA", "300200203", "300200208"],
    "Melão Amarelo": ["MELAO AMARELO", "MELÃO AMARELO"],
    "Melão Gaia": ["MELAO GAIA", "MELÃO GAIA", "MELAO GALIA", "MELÃO GALIA", "3102006"],
    "Melão Cantaloupe": ["CANTALOUPE"], "Tangerina Cumbuca": ["CUMBUCA", "830100903"],
    "Tangerina Ponkan": ["PONKAN"], "Ameixa": ["AMEIXA"],
    "Pêssego": ["PESSEGO", "PÊSSEGO"], "Nectarina": ["NECTARINA"],
    "Morango": ["MORANGO"], "Mirtilo": ["MIRTILO", "BLUEBERRY"], "Abacaxi": ["ABACAXI"],
    "Manga Palmer": ["MANGA PALMER"], "Manga Tommy": ["MANGA TOMMY"], "Laranja": ["LARANJA"],
    "Limão Tahiti": ["LIMAO TAHITI", "LIMÃO TAHITI"],
    "Limão Siciliano": ["LIMAO SICILIANO", "LIMÃO SICILIANO"],
    "Tomate Roma": ["TOMATE ROMA", "ROMA"],
}

RODAPE = [
    "É DE RESPONSABILIDADE DO VENDEDOR:",
    "AVALIAR DIARIAMENTE A QUALIDADE E ARMAZENAGEM DE CADA PRODUTO DE SUA RESPONSABILIDADE.",
    "CONFERIR O QUE ESTA EM CADA PAVILHÃO",
    "CONFERIR O QUE ESTA NA VENDA FUTURA E ACOMPANHAR DIARIAMENTE",
    "CONFERIR O QUE ESTA ARMAZENADO EM OUTROS FRIGORIFICOS",
    "VENDER ATÉ A ÚLTIMA CAIXA",
    "",
    "DEVOLUCAO SO SE FOR NO MESMO DIA",
    "",
    "MERCADORIAS NO SOL",
    "",
    "CAMINHOES REFRIGERADOS SEMPRE FECHADOS",
]

def mapear_produto(desc):
    desc_u = desc.upper()
    for prod_meta, keywords in MAPA_PRODUTO.items():
        for kw in keywords:
            if kw in desc_u:
                return prod_meta
    return None

GITHUB_TOKEN = st.secrets.get("GITHUB_TOKEN", "")
GITHUB_REPO  = st.secrets.get("GITHUB_REPO", "Othilbh/Relatorios-Comerciais")
GITHUB_FILE_METAS    = "data/metas_semana.json"
GITHUB_FILE_PRODUTOS = "data/produtos_extra.json"

def github_get(filepath):
    if not GITHUB_TOKEN: return None, None
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{filepath}"
    r = requests.get(url, headers={"Authorization": f"token {GITHUB_TOKEN}"})
    if r.status_code == 200:
        data = r.json()
        return json.loads(base64.b64decode(data["content"]).decode("utf-8")), data["sha"]
    return None, None

def github_save(filepath, content, sha=None):
    if not GITHUB_TOKEN: return False
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{filepath}"
    body = {"message": f"Atualiza {filepath}",
            "content": base64.b64encode(json.dumps(content, ensure_ascii=False, indent=2).encode()).decode()}
    if sha: body["sha"] = sha
    r = requests.put(url, headers={"Authorization": f"token {GITHUB_TOKEN}"}, json=body)
    return r.status_code in [200, 201]

if "produtos_meta" not in st.session_state:
    dados, _ = github_get(GITHUB_FILE_METAS)
    st.session_state.produtos_meta = dados or []
if "produtos_extra" not in st.session_state:
    dados, _ = github_get(GITHUB_FILE_PRODUTOS)
    st.session_state.produtos_extra = dados or []
if "vendido" not in st.session_state: st.session_state.vendido = {}
if "estoque" not in st.session_state: st.session_state.estoque = {}
if "sem_vendedor" not in st.session_state: st.session_state.sem_vendedor = []

PRODUTOS_LISTA = PRODUTOS_DEFAULT + [p for p in st.session_state.produtos_extra if p not in PRODUTOS_DEFAULT]

COR_H     = "1A3A5C"
COR_S     = "2E6DA4"
COR_CINZA = "D9D9D9"
COR_ALT   = "EBF3FB"
COR_TOTAL = "BDD7EE"
COR_V = "C6EFCE"; COR_VF = "276221"
COR_A = "FFEB9C"; COR_AF = "7D6608"
COR_R = "FFC7CE"; COR_RF = "9C0006"

def _f(c): return PatternFill("solid", fgColor=c)
def _t(bold=False, color="000000", size=9): return Font(bold=bold, color=color, size=size, name="Calibri")
def _b(style="thin", color="BFBFBF"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def _bh():
    m = Side(style="medium", color="1A3A5C")
    return Border(left=m, right=m, top=m, bottom=m)
def _a(h="center", wrap=False): return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def set_a4(ws):
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

def cor(pct):
    if pct >= 100: return COR_V, COR_VF
    if pct >= 50:  return COR_A, COR_AF
    return COR_R, COR_RF

def ch(ws, r, c, v, bg=COR_H, fg="FFFFFF", bold=True, sz=9, h="center"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill = _f(bg); cell.font = _t(bold=bold, color=fg, size=sz)
    cell.alignment = _a(h); cell.border = _b("thin", "1A3A5C")
    return cell

def cd(ws, r, c, v, fmt=None, bg="FFFFFF", bold=False, h="center", color="000000"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill = _f(bg); cell.font = _t(bold=bold, color=color, size=9)
    cell.alignment = _a(h); cell.border = _b()
    if fmt: cell.number_format = fmt
    return cell

def cp(ws, r, c, pct):
    bg_p, fg_p = cor(pct)
    cell = ws.cell(row=r, column=c, value=round(pct, 1))
    cell.fill = _f(bg_p); cell.font = _t(bold=True, color=fg_p, size=9)
    cell.alignment = _a("center"); cell.border = _b("thin", "1A3A5C")
    cell.number_format = '0.0"%"'
    return cell


# ══════════════════════════════════════════════════════════════════════════
# ABA VENDEDOR
# ══════════════════════════════════════════════════════════════════════════
def gerar_aba_vendedor(wb, vendedor, data_ref, itens_estoque, metas, vendido_v):
    ws = wb.create_sheet(vendedor[:31])
    ws.sheet_view.showGridLines = False
    set_a4(ws)

    # ── Cabeçalho ──────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 20
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"Vendedor : {vendedor.upper()}"
    c.font = _t(bold=True, size=11)
    c.alignment = _a("left")
    c.border = _bh()

    ws.row_dimensions[2].height = 13
    ws["A2"] = f"Data {data_ref}"
    ws["A2"].font = _t(size=10)

    ws.row_dimensions[3].height = 8

    # ── Header Estoque ──────────────────────────────────────────────────
    ws.row_dimensions[4].height = 28
    hdrs_e = ["Produto", "Complemento", "Data Entrada", "Saldo Atual",
              "Qtde Vendida", "Custo Unitario", "Md Venda"]
    for ci, h in enumerate(hdrs_e, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.fill = _f(COR_CINZA)
        cell.font = _t(bold=True, color="000000", size=9)
        cell.alignment = _a("center")
        cell.border = _b("thin", "1A3A5C")

    # ── Dados Estoque ───────────────────────────────────────────────────
    row = 5
    for i, item in enumerate(itens_estoque):
        bg = "FFFFFF" if i % 2 == 0 else COR_ALT
        ws.row_dimensions[row].height = 13

        cd(ws, row, 1, item["descricao"], bg=bg, h="left", color="1A3A5C")
        cd(ws, row, 2, item["complemento"], bg=bg, h="center")
        cd(ws, row, 3, item["data_entrada"], bg=bg, h="center")
        cd(ws, row, 4, item["saldo_atual"], bg=bg, h="center")

        c_q = cd(ws, row, 5, item["qtd_vendida"], bg=bg, h="center")
        if item["qtd_vendida"] == 0:
            c_q.font = _t(color="FF0000", size=9)

        c_cu = ws.cell(row=row, column=6, value=item["custo"])
        c_cu.fill = _f(bg)
        c_cu.font = _t(color="FF0000" if item["custo"] == 0 else "000000", size=9)
        c_cu.alignment = _a("right")
        c_cu.border = _b()
        c_cu.number_format = 'R$ #,##0.00'

        c_mv = ws.cell(row=row, column=7, value=item["md_venda"])
        c_mv.fill = _f(bg)
        c_mv.font = _t(color="FF0000" if item["md_venda"] == 0 else "000000", size=9)
        c_mv.alignment = _a("right")
        c_mv.border = _b()
        c_mv.number_format = 'R$ #,##0.00'

        row += 1

    row += 1

    # ── Título Metas ────────────────────────────────────────────────────
    if metas:
        ws.row_dimensions[row].height = 20
        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row=row, column=1,
            value=f"METAS SEMANAIS — {vendedor.upper()} — {data_ref}")
        c.fill = _f(COR_H)
        c.font = _t(bold=True, color="FFFFFF", size=11)
        c.alignment = _a("center")
        c.border = _bh()
        row += 1

        # ── Header Metas ────────────────────────────────────────────────
        ws.row_dimensions[row].height = 15
        hdrs_m = ["Produto", "Meta (cx)", "Vendido (cx)", "Falta (cx)", "%"]
        for ci, h in enumerate(hdrs_m, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.fill = _f(COR_S)
            cell.font = _t(bold=True, color="FFFFFF", size=9)
            cell.alignment = _a("center")
            cell.border = _b("thin", "1A3A5C")
        row += 1

        # ── Linhas Metas ────────────────────────────────────────────────
        total_meta = total_vend = 0
        for item in metas:
            produto = item["produto"]
            est = item["estoque"]
            meta = math.ceil(est * PERCENTUAIS.get(vendedor, 0))
            vend = vendido_v.get(produto, 0.0)
            falta = max(meta - vend, 0)
            pct = (vend / meta * 100) if meta > 0 else 0.0
            total_meta += meta
            total_vend += vend

            ws.row_dimensions[row].height = 13
            cd(ws, row, 1, produto, bg="FFFFFF", h="left")
            cd(ws, row, 2, meta, bg="FFFFFF", h="center")
            c_v = cd(ws, row, 3, round(vend, 1), bg="FFFFFF", h="center")
            if vend == 0:
                c_v.font = _t(color="FF0000", size=9)
            cd(ws, row, 4, round(falta, 1), bg="FFFFFF", h="center")
            cp(ws, row, 5, pct)
            row += 1

        # ── Total Metas ─────────────────────────────────────────────────
        pct_t = (total_vend / total_meta * 100) if total_meta > 0 else 0
        ws.row_dimensions[row].height = 15
        for ci, val in enumerate(["TOTAL", total_meta, round(total_vend, 1),
                                   round(max(total_meta - total_vend, 0), 1), ""], 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.fill = _f(COR_TOTAL)
            c.font = _t(bold=True, size=9)
            c.alignment = _a("left" if ci == 1 else "center")
            c.border = _b("medium", "1A3A5C")
        c_pt = cp(ws, row, 5, pct_t)
        c_pt.border = _b("medium", "1A3A5C")
        c_pt.font = _t(bold=True, color=cor(pct_t)[1], size=9)
        row += 2

        # ── KPIs abaixo das metas ────────────────────────────────────────
        kpis = [
            ("Meta Total (cx)", total_meta, COR_S, "FFFFFF"),
            ("Vendido (cx)", round(total_vend, 1), COR_V, COR_VF),
            ("Falta (cx)", round(max(total_meta - total_vend, 0), 1), COR_R, COR_RF),
            ("% Atingido", f"{pct_t:.1f}%", *cor(pct_t)),
        ]
        ws.row_dimensions[row].height = 13
        ws.row_dimensions[row + 1].height = 24
        for ki, (label, valor, bg_k, fg_k) in enumerate(kpis):
            col_k = ki + 1
            c_l = ws.cell(row=row, column=col_k, value=label)
            c_l.fill = _f(COR_CINZA)
            c_l.font = _t(bold=True, size=8)
            c_l.alignment = _a("center")
            c_l.border = _b("thin", "1A3A5C")
            c_v = ws.cell(row=row + 1, column=col_k, value=valor)
            c_v.fill = _f(bg_k)
            c_v.font = _t(bold=True, color=fg_k, size=14)
            c_v.alignment = _a("center")
            c_v.border = _b("medium", "1A3A5C")
        row += 3

    # ── Rodapé ──────────────────────────────────────────────────────────
    for texto in RODAPE:
        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row=row, column=1, value=texto)
        if texto.startswith("É DE RESPONSABILIDADE"):
            c.font = _t(bold=True, size=10, color="FF0000")
            ws.row_dimensions[row].height = 15
        else:
            c.font = _t(size=9)
            ws.row_dimensions[row].height = 12
        c.alignment = _a("left")
        row += 1

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 13
    return ws


# ══════════════════════════════════════════════════════════════════════════
# ABA RESUMO GERAL
# ══════════════════════════════════════════════════════════════════════════
def gerar_aba_resumo(wb, metas, vendido_todos, data_ref, periodo):
    ws = wb.create_sheet("Resumo Geral")
    ws.sheet_view.showGridLines = False
    set_a4(ws)

    n_cols = 1 + len(VENDEDORES_ATIVOS) * 3 + 3

    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value = f"OTHIL — RESUMO GERAL DE METAS  |  {periodo}  |  {data_ref}"
    c.fill = _f(COR_H); c.font = _t(bold=True, color="FFFFFF", size=12)
    c.alignment = _a("center"); c.border = _bh()
    ws.row_dimensions[1].height = 22

    ws.row_dimensions[2].height = 16
    ws.cell(row=2, column=1).border = _b("medium", "1A3A5C")
    col = 2
    for v in VENDEDORES_ATIVOS:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 2)
        c = ws.cell(row=2, column=col, value=v)
        c.fill = _f(COR_S); c.font = _t(bold=True, color="FFFFFF", size=9)
        c.alignment = _a("center"); c.border = _b("medium", "1A3A5C")
        col += 3
    ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 2)
    c = ws.cell(row=2, column=col, value="TOTAL GERAL")
    c.fill = _f(COR_H); c.font = _t(bold=True, color="FFFFFF", size=9)
    c.alignment = _a("center"); c.border = _b("medium", "1A3A5C")

    ws.row_dimensions[3].height = 13
    ch(ws, 3, 1, "Produto", bg=COR_CINZA, fg="000000", sz=9, h="left")
    col = 2
    for _ in VENDEDORES_ATIVOS:
        for h in ["Meta", "Vend", "%"]:
            ch(ws, 3, col, h, bg=COR_CINZA, fg="000000", sz=8)
            col += 1
    for h in ["Meta", "Vend", "%"]:
        ch(ws, 3, col, h, bg=COR_CINZA, fg="000000", sz=8)
        col += 1

    row = 4
    tot_meta_v = {v: 0 for v in VENDEDORES_ATIVOS}
    tot_vend_v = {v: 0.0 for v in VENDEDORES_ATIVOS}

    for i, item in enumerate(metas):
        produto = item["produto"]
        est = item["estoque"]
        bg = "FFFFFF" if i % 2 == 0 else COR_ALT
        ws.row_dimensions[row].height = 13

        c = ws.cell(row=row, column=1, value=produto)
        c.fill = _f(bg); c.font = _t(size=9, color="1A3A5C")
        c.alignment = _a("left"); c.border = _b()

        col = 2
        tot_meta_l = tot_vend_l = 0
        for v in VENDEDORES_ATIVOS:
            meta_v = math.ceil(est * PERCENTUAIS.get(v, 0))
            vend_v = round(vendido_todos.get(v, {}).get(produto, 0.0), 1)
            pct_v  = (vend_v / meta_v * 100) if meta_v > 0 else 0.0
            tot_meta_v[v] += meta_v; tot_vend_v[v] += vend_v
            tot_meta_l += meta_v; tot_vend_l += vend_v
            cd(ws, row, col, meta_v, bg=bg, h="center"); col += 1
            c_vv = cd(ws, row, col, vend_v, bg=bg, h="center")
            if vend_v == 0: c_vv.font = _t(color="FF0000", size=9)
            col += 1
            cp(ws, row, col, pct_v); col += 1

        pct_l = (tot_vend_l / tot_meta_l * 100) if tot_meta_l > 0 else 0
        cd(ws, row, col, tot_meta_l, bg=COR_TOTAL, bold=True, h="center"); col += 1
        cd(ws, row, col, round(tot_vend_l, 1), bg=COR_TOTAL, bold=True, h="center"); col += 1
        cp(ws, row, col, pct_l)
        row += 1

    ws.row_dimensions[row].height = 15
    c = ws.cell(row=row, column=1, value="TOTAL")
    c.fill = _f(COR_H); c.font = _t(bold=True, color="FFFFFF", size=9)
    c.alignment = _a("left"); c.border = _b("medium", "1A3A5C")

    col = 2; grand_m = grand_v = 0
    for v in VENDEDORES_ATIVOS:
        pct_v = (tot_vend_v[v] / tot_meta_v[v] * 100) if tot_meta_v[v] > 0 else 0
        grand_m += tot_meta_v[v]; grand_v += tot_vend_v[v]
        for val in [tot_meta_v[v], round(tot_vend_v[v], 1)]:
            c = ws.cell(row=row, column=col, value=val)
            c.fill = _f(COR_TOTAL); c.font = _t(bold=True, size=9)
            c.alignment = _a("center"); c.border = _b("thin", "1A3A5C")
            col += 1
        c_p = cp(ws, row, col, pct_v)
        c_p.border = _b("medium", "1A3A5C")
        col += 1

    pct_grand = (grand_v / grand_m * 100) if grand_m > 0 else 0
    for val in [grand_m, round(grand_v, 1)]:
        c = ws.cell(row=row, column=col, value=val)
        c.fill = _f(COR_H); c.font = _t(bold=True, color="FFFFFF", size=9)
        c.alignment = _a("center"); c.border = _b("medium", "1A3A5C")
        col += 1
    c_gp = ws.cell(row=row, column=col, value=round(pct_grand, 1))
    bg_gp, fg_gp = cor(pct_grand)
    c_gp.fill = _f(bg_gp); c_gp.font = _t(bold=True, color=fg_gp, size=9)
    c_gp.alignment = _a("center"); c_gp.border = _b("medium", "1A3A5C")
    c_gp.number_format = '0.0"%"'

    ws.column_dimensions["A"].width = 24
    col = 2
    for _ in VENDEDORES_ATIVOS:
        ws.column_dimensions[get_column_letter(col)].width = 7
        ws.column_dimensions[get_column_letter(col + 1)].width = 7
        ws.column_dimensions[get_column_letter(col + 2)].width = 6
        col += 3
    for extra in range(3):
        ws.column_dimensions[get_column_letter(col + extra)].width = 8
    return ws


# ══════════════════════════════════════════════════════════════════════════
# ABA DASHBOARD
# ══════════════════════════════════════════════════════════════════════════
def gerar_aba_dashboard(wb, metas, vendido_todos, data_ref, periodo):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"OTHIL — DASHBOARD DE METAS  |  {periodo}"
    c.fill = _f(COR_H); c.font = _t(bold=True, color="FFFFFF", size=13)
    c.alignment = _a("center"); c.border = _bh()
    ws.row_dimensions[1].height = 26

    total_meta_g = sum(math.ceil(item["estoque"] * PERCENTUAIS.get(v, 0))
                       for item in metas for v in VENDEDORES_ATIVOS)
    total_vend_g = sum(vendido_todos.get(v, {}).get(item["produto"], 0)
                       for item in metas for v in VENDEDORES_ATIVOS)
    pct_g = (total_vend_g / total_meta_g * 100) if total_meta_g > 0 else 0
    n_criticos = sum(1 for item in metas
        if (sum(vendido_todos.get(v, {}).get(item["produto"], 0) for v in VENDEDORES_ATIVOS) /
            max(sum(math.ceil(item["estoque"] * PERCENTUAIS.get(v, 0)) for v in VENDEDORES_ATIVOS), 1) * 100) < 50)

    kpis_g = [
        ("Meta Total Geral (cx)", total_meta_g, COR_S, "FFFFFF"),
        ("Vendido Total (cx)", round(total_vend_g, 1), COR_V, COR_VF),
        ("Falta Total (cx)", round(max(total_meta_g - total_vend_g, 0), 1), COR_R, COR_RF),
        ("% Geral Atingido", f"{pct_g:.1f}%", *cor(pct_g)),
        ("Produtos Críticos <50%", n_criticos, "C62828", "FFFFFF"),
        ("Total Produtos", len(metas), COR_H, "FFFFFF"),
    ]

    ws.row_dimensions[2].height = 13
    ws.row_dimensions[3].height = 26
    for ki, (label, valor, bg_k, fg_k) in enumerate(kpis_g):
        col_k = ki + 1
        c_l = ws.cell(row=2, column=col_k, value=label)
        c_l.fill = _f(COR_CINZA); c_l.font = _t(bold=True, size=8)
        c_l.alignment = _a("center"); c_l.border = _b("thin", "1A3A5C")
        c_v = ws.cell(row=3, column=col_k, value=valor)
        c_v.fill = _f(bg_k); c_v.font = _t(bold=True, color=fg_k, size=14)
        c_v.alignment = _a("center"); c_v.border = _b("medium", "1A3A5C")

    row = 5

    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1, value="RANKING DE VENDEDORES")
    c.fill = _f(COR_S); c.font = _t(bold=True, color="FFFFFF", size=10)
    c.alignment = _a("center"); ws.row_dimensions[row].height = 16
    row += 1

    for ci, h in enumerate(["#", "Vendedor", "Meta (cx)", "Vendido (cx)", "Falta (cx)", "% Atingido", "Status", "% da Meta"], 1):
        ch(ws, row, ci, h, bg=COR_CINZA, fg="000000", sz=9)
    ws.row_dimensions[row].height = 14
    row += 1

    ranking = []
    for v in VENDEDORES_ATIVOS:
        meta_v = sum(math.ceil(item["estoque"] * PERCENTUAIS.get(v, 0)) for item in metas)
        vend_v = sum(vendido_todos.get(v, {}).get(item["produto"], 0) for item in metas)
        pct_v  = (vend_v / meta_v * 100) if meta_v > 0 else 0
        ranking.append((v, meta_v, round(vend_v, 1), max(meta_v - vend_v, 0), round(pct_v, 1)))
    ranking.sort(key=lambda x: x[4], reverse=True)

    for pos, (v, meta_v, vend_v, falta_v, pct_v) in enumerate(ranking, 1):
        bg_r, fg_r = cor(pct_v)
        status = "✅ Atingiu" if pct_v >= 100 else "⚠️ Andamento" if pct_v >= 50 else "❌ Abaixo"
        medal = "1°" if pos == 1 else "2°" if pos == 2 else "3°" if pos == 3 else f"{pos}°"
        ws.row_dimensions[row].height = 14
        cd(ws, row, 1, medal, h="center")
        cd(ws, row, 2, v, h="left", color="1A3A5C")
        cd(ws, row, 3, meta_v, h="center")
        c_vv = cd(ws, row, 4, vend_v, h="center")
        if vend_v == 0: c_vv.font = _t(color="FF0000", size=9)
        cd(ws, row, 5, round(falta_v, 1), h="center")
        cp(ws, row, 6, pct_v)
        cd(ws, row, 7, status, bg=bg_r, h="center", color=fg_r)
        cd(ws, row, 8, f"{int(PERCENTUAIS.get(v, 0) * 100)}%", h="center", color="666666")
        row += 1

    row += 1

    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1, value="PRODUTOS CRÍTICOS — ABAIXO DE 50%")
    c.fill = _f("C62828"); c.font = _t(bold=True, color="FFFFFF", size=10)
    c.alignment = _a("center"); ws.row_dimensions[row].height = 16
    row += 1

    for ci, h in enumerate(["Produto", "Meta Total", "Vendido Total", "Falta", "% Geral", "Melhor Vendedor", "", ""], 1):
        ch(ws, row, ci, h, bg=COR_CINZA, fg="000000", sz=9)
    ws.row_dimensions[row].height = 13
    row += 1

    criticos = []
    for item in metas:
        produto = item["produto"]; est = item["estoque"]
        meta_t = sum(math.ceil(est * PERCENTUAIS.get(v, 0)) for v in VENDEDORES_ATIVOS)
        vend_t = sum(vendido_todos.get(v, {}).get(produto, 0) for v in VENDEDORES_ATIVOS)
        pct_t  = (vend_t / meta_t * 100) if meta_t > 0 else 0
        if pct_t < 50:
            melhor_v = max(VENDEDORES_ATIVOS,
                key=lambda v: (vendido_todos.get(v, {}).get(produto, 0) /
                    max(math.ceil(est * PERCENTUAIS.get(v, 0)), 1) * 100))
            criticos.append((produto, meta_t, round(vend_t, 1),
                             round(max(meta_t - vend_t, 0), 1), round(pct_t, 1), melhor_v))
    criticos.sort(key=lambda x: x[4])

    if not criticos:
        ws.merge_cells(f"A{row}:H{row}")
        c = ws.cell(row=row, column=1, value="Nenhum produto crítico! Todos acima de 50%.")
        c.fill = _f(COR_V); c.font = _t(bold=True, color=COR_VF, size=10)
        c.alignment = _a("center"); ws.row_dimensions[row].height = 18
    else:
        for produto, meta_t, vend_t, falta_t, pct_t, melhor_v in criticos:
            ws.row_dimensions[row].height = 13
            cd(ws, row, 1, produto, bg=COR_R, h="left", color=COR_RF)
            cd(ws, row, 2, meta_t, bg=COR_R, h="center", color=COR_RF)
            cd(ws, row, 3, vend_t, bg=COR_R, h="center", color=COR_RF)
            cd(ws, row, 4, falta_t, bg=COR_R, h="center", color=COR_RF)
            cp(ws, row, 5, pct_t)
            cd(ws, row, 6, melhor_v, bg=COR_V, h="center", color=COR_VF)
            cd(ws, row, 7, ""); cd(ws, row, 8, "")
            row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 11
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 10
    return ws


# ══════════════════════════════════════════════════════════════════════════
# INTERFACE
# ══════════════════════════════════════════════════════════════════════════
st.subheader("1️⃣ Metas da Semana")
col_ini, col_fim = st.columns(2)
with col_ini:
    semana_ini = st.date_input("Início", value=date.today() - timedelta(days=date.today().weekday()))
with col_fim:
    semana_fim = st.date_input("Fim", value=semana_ini + timedelta(days=5))
periodo = f"{semana_ini.strftime('%d/%m/%Y')} a {semana_fim.strftime('%d/%m/%Y')}"

st.caption("Busque ou digite um produto para adicionar:")
produtos_ja = [p["produto"] for p in st.session_state.produtos_meta]
produtos_disp = [p for p in PRODUTOS_LISTA if p not in produtos_ja]

col_b, col_n, col_e, col_btn = st.columns([2, 2, 1, 1])
with col_b:
    prod_sel = st.selectbox("Buscar", options=[""] + produtos_disp,
        format_func=lambda x: "🔍 Selecione da lista..." if x == "" else x,
        label_visibility="collapsed")
with col_n:
    prod_novo = st.text_input("Novo", placeholder="✏️ Ou digite produto novo...",
        label_visibility="collapsed")
with col_e:
    est_input = st.number_input("Est", min_value=0.0, step=1.0,
        label_visibility="collapsed", placeholder="Estoque CX")
with col_btn:
    if st.button("➕ Adicionar", use_container_width=True, type="primary"):
        prod_final = prod_novo.strip() if prod_novo.strip() else prod_sel
        if prod_final and prod_final != "":
            if prod_final not in produtos_ja:
                st.session_state.produtos_meta.append({"produto": prod_final, "estoque": est_input})
                if prod_final not in PRODUTOS_LISTA:
                    st.session_state.produtos_extra.append(prod_final)
                    _, sha = github_get(GITHUB_FILE_PRODUTOS)
                    github_save(GITHUB_FILE_PRODUTOS, st.session_state.produtos_extra, sha)
                    st.toast(f"✅ '{prod_final}' salvo na lista permanente!")
                st.rerun()
        else:
            st.warning("Selecione ou digite um produto.")

if st.session_state.produtos_meta:
    rows = []
    for item in st.session_state.produtos_meta:
        est = item["estoque"]
        row = {"Produto": item["produto"], "Estoque CX": int(est)}
        for v in VENDEDORES_ATIVOS:
            row[f"{v} ({int(PERCENTUAIS[v] * 100)}%)"] = math.ceil(est * PERCENTUAIS[v])
        rows.append(row)

    df_edit = st.data_editor(pd.DataFrame(rows), use_container_width=True, hide_index=True,
        disabled=["Produto"] + [f"{v} ({int(PERCENTUAIS[v] * 100)}%)" for v in VENDEDORES_ATIVOS],
        key="editor_metas")
    for i, row in df_edit.iterrows():
        if i < len(st.session_state.produtos_meta):
            st.session_state.produtos_meta[i]["estoque"] = row["Estoque CX"]

    col_s, col_a, col_i = st.columns([1, 1, 3])
    with col_s:
        if st.button("💾 Salvar Metas", type="primary", use_container_width=True):
            _, sha = github_get(GITHUB_FILE_METAS)
            ok = github_save(GITHUB_FILE_METAS, st.session_state.produtos_meta, sha)
            st.success("✅ Metas salvas!") if ok else st.error("❌ Erro ao salvar.")
    with col_a:
        if st.button("🗑️ Apagar Metas", type="secondary", use_container_width=True):
            _, sha = github_get(GITHUB_FILE_METAS)
            github_save(GITHUB_FILE_METAS, [], sha)
            st.session_state.produtos_meta = []
            st.session_state.vendido = {}
            st.session_state.estoque = {}
            st.rerun()
    with col_i:
        st.caption(f"💡 Semana: **{periodo}**")

st.divider()
st.subheader("2️⃣ Upload dos PDFs")
col_p1, col_p2 = st.columns(2)
with col_p1:
    st.caption("📊 PDF de Vendas Acumuladas")
    pdf_vendas = st.file_uploader("Vendas", type=["pdf"], label_visibility="collapsed")
with col_p2:
    st.caption("📦 PDF de Estoque Físico")
    pdf_estoque = st.file_uploader("Estoque", type=["pdf"], label_visibility="collapsed")

if pdf_vendas:
    with st.spinner("🔍 Lendo vendas..."):
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(pdf_vendas.read()); tmp_path = tmp.name
        vendas_raw = extrair_vendas_por_vendedor(tmp_path)
        os.unlink(tmp_path)
    vendido_c = {v: {} for v in VENDEDORES_ATIVOS}
    for vend, prods in vendas_raw.items():
        if vend not in vendido_c: continue
        for desc, qtd in prods.items():
            pm = mapear_produto(desc)
            if pm: vendido_c[vend][pm] = vendido_c[vend].get(pm, 0) + qtd
    st.session_state.vendido = vendido_c
    found = [v for v in vendas_raw if v in VENDEDORES_ATIVOS]
    st.success(f"✅ Vendas: {', '.join(found)}")

if pdf_estoque:
    with st.spinner("🔍 Lendo estoque..."):
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(pdf_estoque.read()); tmp_path = tmp.name
        est_raw = extrair_estoque_por_vendedor(tmp_path)
        os.unlink(tmp_path)
    st.session_state.estoque = {v: est_raw.get(v, []) for v in VENDEDORES_ATIVOS}
    st.session_state.sem_vendedor = est_raw.get("SEM_VENDEDOR", [])
    total = sum(len(v) for v in st.session_state.estoque.values())
    st.success(f"✅ Estoque: {total} produto(s) distribuídos.")
    if st.session_state.sem_vendedor:
        with st.expander(f"⚠️ {len(st.session_state.sem_vendedor)} produto(s) SEM VENDEDOR", expanded=True):
            df_sv = pd.DataFrame(st.session_state.sem_vendedor)[["codigo", "descricao", "data_entrada", "saldo_atual"]]
            df_sv.columns = ["Código", "Descrição", "Data Entrada", "Saldo Atual"]
            st.dataframe(df_sv, use_container_width=True, hide_index=True)

st.divider()
st.subheader("3️⃣ Gerar Relatórios")

if st.button("📋 Gerar Relatórios Completos", use_container_width=True, type="primary"):
    if not st.session_state.estoque:
        st.warning("Faça upload do PDF de Estoque primeiro.")
    else:
        with st.spinner("⚙️ Gerando Excel..."):
            wb = Workbook()
            wb.remove(wb.active)
            data_ref = date.today().strftime("%d/%m/%Y")

            if st.session_state.produtos_meta:
                gerar_aba_dashboard(wb, st.session_state.produtos_meta,
                    st.session_state.vendido, data_ref, periodo)
                gerar_aba_resumo(wb, st.session_state.produtos_meta,
                    st.session_state.vendido, data_ref, periodo)

            for vendedor in VENDEDORES_ATIVOS:
                gerar_aba_vendedor(wb, vendedor, data_ref,
                    st.session_state.estoque.get(vendedor, []),
                    st.session_state.produtos_meta,
                    st.session_state.vendido.get(vendedor, {}))

            buf = io.BytesIO()
            wb.save(buf); buf.seek(0)

        nome = f"Relatorios_{date.today().strftime('%d%m%Y')}.xlsx"
        st.success("✅ Relatórios gerados!")
        st.download_button(
            label="⬇️ Baixar Excel Completo",
            data=buf, file_name=nome,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )